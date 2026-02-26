from datetime import date, timedelta, datetime, time
import uuid
from typing import Any, Dict, List

import os
from fastapi import HTTPException, status, Depends
from sqlalchemy.orm import Session

# 👇 Ajusta estos imports a tu proyecto, tal y como ya los uses en tu API
from app.database import get_db               # Debe existir en tu proyecto
from app.models import Usuario                # Modelo de usuario (ajusta el nombre si difiere)
from app.security import hash_password        # Si tienes util para hashear. Si no, te doy una alternativa abajo.

import json
import io
import smtplib
from email.message import EmailMessage
from zoneinfo import ZoneInfo

from fastapi import (
    FastAPI,
    Depends,
    HTTPException,
    UploadFile,
    File,
    Request,
)
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from openpyxl import load_workbook, Workbook
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import requests

from . import models, schemas
from .database import SessionLocal, engine

models.Base.metadata.create_all(bind=engine)

app = FastAPI(
    title="Reserva de aulas - IES Antonio García Bellido",
    version="0.4.0",
)

# Archivos estáticos (logo, CSS, etc.)
app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")

security = HTTPBearer()

TEAMS_WEBHOOK_URL = os.getenv("TEAMS_WEBHOOK_URL")
APP_TZ = ZoneInfo(os.getenv("APP_TZ", "Europe/Madrid"))
SMTP_HOST = os.getenv("SMTP_HOST")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
SMTP_FROM = os.getenv("SMTP_FROM")

scheduler: AsyncIOScheduler | None = None


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def generar_token() -> str:
    return uuid.uuid4().hex


def enviar_mensaje_teams(texto: str) -> None:
    """
    Envía un mensaje sencillo a un webhook de Teams, si está configurado.
    """
    if not TEAMS_WEBHOOK_URL:
        return
    try:
        requests.post(TEAMS_WEBHOOK_URL, json={"text": texto}, timeout=5)
    except Exception:
        # En caso de error con Teams, no se rompe la API principal
        return


def enviar_correo(
    asunto: str,
    cuerpo: str,
    destinatarios: list[str],
    adjuntos: list[tuple[str, bytes, str]] | None = None,
) -> None:
    """
    Envía un correo con adjuntos opcionales.
    adjuntos: lista de tuplas (nombre_fichero, contenido_bytes, mime_type)
    """
    if not SMTP_HOST or not SMTP_FROM or not destinatarios:
        return

    try:
        msg = EmailMessage()
        msg["Subject"] = asunto
        msg["From"] = SMTP_FROM
        msg["To"] = ", ".join(destinatarios)
        msg.set_content(cuerpo)

        for nombre, contenido, mime_type in adjuntos or []:
            maintype, subtype = mime_type.split("/", 1)
            msg.add_attachment(
                contenido,
                maintype=maintype,
                subtype=subtype,
                filename=nombre,
            )

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as server:
            server.starttls()
            if SMTP_USER and SMTP_PASSWORD:
                server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
    except Exception:
        # No rompemos la app principal si falla el correo
        return


def hora_inicio_franja(franja: schemas.FranjaHorariaEnum) -> time:
    """
    Convierte una franja "HH:MM-HH:MM" en hora de inicio (time).
    """
    try:
        inicio_str = franja.value.split("-", 1)[0]
        return time.fromisoformat(inicio_str)
    except Exception:
        # Fallback muy defensivo: si algo raro llega, bloqueamos por seguridad en validaciones.
        return time(23, 59)


def registrar_auditoria(
    db: Session,
    *,
    actor: models.Usuario | None,
    accion: str,
    entidad: str,
    entidad_id: int | None = None,
    detalle: dict[str, Any] | None = None,
    request: Request | None = None,
) -> None:
    """
    Registra una fila de auditoría. Si falla, no debe romper la operación principal.
    """
    try:
        actor_ip = None
        if request and request.client:
            actor_ip = request.client.host

        fila = models.Auditoria(
            actor_usuario_id=getattr(actor, "id", None) if actor else None,
            actor_nombre=getattr(actor, "nombre", None) if actor else None,
            actor_rol=getattr(actor, "rol", None) if actor else None,
            actor_email=getattr(actor, "email", None) if actor else None,
            actor_ip=actor_ip,
            accion=accion,
            entidad=entidad,
            entidad_id=entidad_id,
            detalle=json.dumps(detalle, ensure_ascii=False) if detalle is not None else None,
        )
        db.add(fila)
    except Exception:
        return


def generar_cuadrante_rango(
    db: Session,
    fecha_inicio: date,
    fecha_fin: date,
    aula: schemas.AulaEnum | None = None,
) -> Dict[str, Any]:
    """
    Genera una estructura de cuadrante entre dos fechas (inclusive).
    Similar a /cuadrante-semanal pero para un rango arbitrario.
    """
    if fecha_inicio > fecha_fin:
        fecha_inicio, fecha_fin = fecha_fin, fecha_inicio

    franjas = [f.value for f in schemas.FranjaHorariaEnum]
    dias = []
    dia = fecha_inicio
    while dia <= fecha_fin:
        dias.append(dia)
        dia += timedelta(days=1)

    aulas_objetivo: List[str]
    if aula:
        aulas_objetivo = [aula.value]
    else:
        aulas_objetivo = [a.value for a in schemas.AulaEnum]

    cuadrante: Dict[str, Dict[str, Dict[str, Any | None]]] = {}
    for nombre_aula in aulas_objetivo:
        cuadrante[nombre_aula] = {}
        for fr in franjas:
            cuadrante[nombre_aula][fr] = {}
            for d in dias:
                cuadrante[nombre_aula][fr][d.isoformat()] = None

    reservas = (
        db.query(models.Reserva)
        .filter(models.Reserva.fecha >= fecha_inicio, models.Reserva.fecha <= fecha_fin)
        .all()
    )

    for r in reservas:
        if r.aula not in aulas_objetivo:
            continue
        if r.franja_horaria not in franjas:
            continue
        dia_str = r.fecha.isoformat()
        if dia_str not in cuadrante[r.aula][r.franja_horaria]:
            continue
        cuadrante[r.aula][r.franja_horaria][dia_str] = {
            "id": r.id,
            "profesor": r.profesor,
        }

    return {
        "fecha_inicio": fecha_inicio.isoformat(),
        "fecha_fin": fecha_fin.isoformat(),
        "dias": [d.isoformat() for d in dias],
        "franjas_horarias": franjas,
        "aulas": aulas_objetivo,
        "cuadrante": cuadrante,
    }


def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db),
) -> models.Usuario:
    token = credentials.credentials
    usuario = db.query(models.Usuario).filter(models.Usuario.api_token == token).first()
    if not usuario:
        raise HTTPException(status_code=401, detail="Token inválido")
    if not usuario.activo:
        raise HTTPException(status_code=403, detail="Usuario inactivo")
    return usuario


def get_admin_user(
    current_user: models.Usuario = Depends(get_current_user),
) -> models.Usuario:
    if current_user.rol != schemas.RolEnum.ADMIN.value:
        raise HTTPException(status_code=403, detail="Solo administradores")
    return current_user


@app.get("/")
def read_root():
    return {
        "mensaje": "API de reservas de aulas IES Antonio García Bellido",
        "aulas_disponibles": [a.value for a in schemas.AulaEnum],
        "franjas_horarias": [f.value for f in schemas.FranjaHorariaEnum],
    }


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    """
    Página de inicio de sesión.
    """
    return templates.TemplateResponse("login.html", {"request": request})


@app.get("/web", response_class=HTMLResponse)
def web_home(request: Request):
    """
    Página principal web con logo y diseño responsive.
    """
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/setup/crear-admin-inicial", response_model=schemas.UsuarioOut)
def crear_admin_inicial(
    datos: schemas.UsuarioCreate,
    request: Request,
    db: Session = Depends(get_db),
):
    """
    Crea el primer usuario administrador.
    Solo se permite si todavía no hay usuarios en la base de datos.
    """
    hay_usuarios = db.query(models.Usuario).first()
    if hay_usuarios:
        raise HTTPException(
            status_code=400,
            detail="Ya existe al menos un usuario. Esta ruta solo es para el primer admin.",
        )

    if datos.rol != schemas.RolEnum.ADMIN:
        raise HTTPException(
            status_code=400,
            detail="El primer usuario debe ser administrador",
        )

    token = generar_token()
    usuario = models.Usuario(
        nombre=datos.nombre,
        email=datos.email,
        rol=datos.rol.value,
        password=datos.password,
        api_token=token,
        activo=datos.activo,
    )
    db.add(usuario)
    db.flush()
    registrar_auditoria(
        db,
        actor=usuario,
        accion="usuario.crear_admin_inicial",
        entidad="Usuario",
        entidad_id=usuario.id,
        detalle={"nombre": usuario.nombre, "email": usuario.email, "rol": usuario.rol},
        request=request,
    )
    db.commit()
    db.refresh(usuario)
    return usuario


@app.post("/auth/login", response_model=schemas.TokenOut)
def login(datos: schemas.LoginRequest, db: Session = Depends(get_db)):
    usuario = db.query(models.Usuario).filter(models.Usuario.email == datos.email).first()
    if not usuario or usuario.password != datos.password:
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    if not usuario.activo:
        raise HTTPException(
            status_code=403,
            detail="Usuario inactivo. Póngase en contacto con el administrador.",
        )

    # Para simplificar el ejemplo, reutilizamos siempre el mismo token
    return schemas.TokenOut(
        token=usuario.api_token,
        rol=schemas.RolEnum(usuario.rol),
        nombre=usuario.nombre,
        email=usuario.email,
    )


@app.post("/usuarios", response_model=schemas.UsuarioOut)
def crear_usuario(
    datos: schemas.UsuarioCreate,
    request: Request,
    admin: models.Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    """
    Crear usuarios (profesores o admins). Solo puede hacerlo un administrador.
    """
    existente = db.query(models.Usuario).filter(
        (models.Usuario.email == datos.email) | (models.Usuario.nombre == datos.nombre)
    ).first()
    if existente:
        raise HTTPException(
            status_code=400,
            detail="Ya existe un usuario con ese nombre o email",
        )

    token = generar_token()
    usuario = models.Usuario(
        nombre=datos.nombre,
        email=datos.email,
        rol=datos.rol.value,
        password=datos.password,
        api_token=token,
        activo=datos.activo,
    )
    db.add(usuario)
    db.flush()
    registrar_auditoria(
        db,
        actor=admin,
        accion="usuario.crear",
        entidad="Usuario",
        entidad_id=usuario.id,
        detalle={"nombre": usuario.nombre, "email": usuario.email, "rol": usuario.rol, "activo": usuario.activo},
        request=request,
    )
    db.commit()
    db.refresh(usuario)
    return usuario


@app.post("/usuarios/importar-profesores")
async def importar_profesores_desde_excel(
    request: Request,
    fichero: UploadFile = File(...),
    admin: models.Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    """
    Importa profesores desde un Excel llamado 'profesores' con dos columnas:
    nombre y email. Se crean usuarios con rol 'profesor'.
    """
    contenido = await fichero.read()
    wb = load_workbook(filename=io.BytesIO(contenido))
    hoja = wb.active

    creados = 0
    ya_existian = 0

    # Suponemos fila 1 como cabecera
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if not fila or all(c is None for c in fila):
            continue
        nombre, email = fila[0], fila[1]
        if not nombre or not email:
            continue

        existe = db.query(models.Usuario).filter(
            (models.Usuario.email == email) | (models.Usuario.nombre == nombre)
        ).first()
        if existe:
            ya_existian += 1
            continue

        token = generar_token()
        usuario = models.Usuario(
            nombre=nombre,
            email=email,
            rol=schemas.RolEnum.PROFESOR.value,
            password="cambiar123",  # contraseña genérica, recomendable cambiarla
            api_token=token,
            activo=True,
        )
        db.add(usuario)
        creados += 1

    db.commit()

    registrar_auditoria(
        db,
        actor=admin,
        accion="usuario.importar_profesores_excel",
        entidad="Usuario",
        entidad_id=None,
        detalle={"creados": creados, "ya_existian": ya_existian, "filename": fichero.filename},
        request=request,
    )
    db.commit()

    return {"creados": creados, "ya_existian": ya_existian}


@app.get("/usuarios", response_model=List[schemas.UsuarioOut])
def listar_usuarios(
    admin: models.Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    """
    Lista todos los usuarios (solo administradores).
    """
    return db.query(models.Usuario).all()


@app.get("/auditoria", response_model=schemas.AuditoriaGroupedOut)
def listar_auditoria(
    limit: int = 200,
    offset: int = 0,
    admin: models.Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    """
    Lista el registro de auditoría (solo administradores), agrupado en dos bloques:
    - reservas: altas/bajas/modificaciones de reservas.
    - otros: operaciones sobre usuarios, auth, etc.
    """
    limite = max(1, min(limit, 500))
    salto = max(0, offset)
    filas = (
        db.query(models.Auditoria)
        .order_by(models.Auditoria.creado_en.desc())
        .offset(salto)
        .limit(limite)
        .all()
    )

    reservas: list[models.Auditoria] = []
    otros: list[models.Auditoria] = []

    for fila in filas:
        # Cualquier acción sobre la entidad "Reserva" se considera del bloque de reservas
        # (crear, borrar, recurrentes, etc.). El resto va a "otros".
        if fila.entidad.lower() == "reserva":
            reservas.append(fila)
        else:
            otros.append(fila)

    return {"reservas": reservas, "otros": otros}


@app.patch("/usuarios/{usuario_id}/activar", response_model=schemas.UsuarioOut)
def activar_usuario(
    usuario_id: int,
    request: Request,
    admin: models.Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    usuario = db.query(models.Usuario).get(usuario_id)
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    usuario.activo = True
    registrar_auditoria(
        db,
        actor=admin,
        accion="usuario.activar",
        entidad="Usuario",
        entidad_id=usuario.id,
        detalle={"email": usuario.email, "nombre": usuario.nombre},
        request=request,
    )
    db.commit()
    db.refresh(usuario)
    return usuario


@app.patch("/usuarios/{usuario_id}/desactivar", response_model=schemas.UsuarioOut)
def desactivar_usuario(
    usuario_id: int,
    request: Request,
    admin: models.Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    usuario = db.query(models.Usuario).get(usuario_id)
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    usuario.activo = False
    registrar_auditoria(
        db,
        actor=admin,
        accion="usuario.desactivar",
        entidad="Usuario",
        entidad_id=usuario.id,
        detalle={"email": usuario.email, "nombre": usuario.nombre},
        request=request,
    )
    db.commit()
    db.refresh(usuario)
    return usuario


@app.post("/usuarios/{usuario_id}/reset-password", response_model=schemas.ResetPasswordOut)
def reset_password_usuario(
    usuario_id: int,
    request: Request,
    admin: models.Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    """
    El administrador puede borrar/renovar la contraseña de un usuario
    cuando la haya olvidado. Se establece una contraseña genérica
    (por ejemplo 'cambiar123') que luego debería cambiarse.
    """
    usuario = db.query(models.Usuario).get(usuario_id)
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    nueva_password = "cambiar123"
    usuario.password = nueva_password
    registrar_auditoria(
        db,
        actor=admin,
        accion="usuario.reset_password",
        entidad="Usuario",
        entidad_id=usuario.id,
        detalle={"email": usuario.email, "nombre": usuario.nombre},
        request=request,
    )
    db.commit()

    return schemas.ResetPasswordOut(
        id=usuario.id,
        email=usuario.email,
        nueva_password=nueva_password,
    )


@app.patch("/usuarios/{usuario_id}/rol", response_model=schemas.UsuarioOut)
def cambiar_rol_usuario(
    usuario_id: int,
    datos: schemas.UsuarioRolUpdate,
    request: Request,
    admin: models.Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    """
    Cambia el rol de un usuario (admin/profesor). Solo administradores.
    """
    usuario = db.query(models.Usuario).get(usuario_id)
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    rol_anterior = usuario.rol
    usuario.rol = datos.rol.value

    registrar_auditoria(
        db,
        actor=admin,
        accion="usuario.cambiar_rol",
        entidad="Usuario",
        entidad_id=usuario.id,
        detalle={
            "email": usuario.email,
            "nombre": usuario.nombre,
            "rol_anterior": rol_anterior,
            "rol_nuevo": usuario.rol,
        },
        request=request,
    )
    db.commit()
    db.refresh(usuario)
    return usuario


@app.delete("/usuarios/{usuario_id}", status_code=204)
def borrar_usuario(
    usuario_id: int,
    request: Request,
    admin: models.Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    """
    Borra un usuario. Solo administradores.
    """
    usuario = db.query(models.Usuario).get(usuario_id)
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    registrar_auditoria(
        db,
        actor=admin,
        accion="usuario.borrar",
        entidad="Usuario",
        entidad_id=usuario.id,
        detalle={"email": usuario.email, "nombre": usuario.nombre, "rol": usuario.rol},
        request=request,
    )
    db.delete(usuario)
    db.commit()

    return


@app.post("/auth/cambiar-password")
def cambiar_password(
    datos: schemas.ChangePasswordRequest,
    request: Request,
    current_user: models.Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Permite que un usuario (admin o profesor) cambie su propia contraseña.
    Debe indicar la contraseña actual y la nueva.
    """
    if current_user.password != datos.password_actual:
        raise HTTPException(status_code=400, detail="La contraseña actual no es correcta")

    current_user.password = datos.password_nueva
    registrar_auditoria(
        db,
        actor=current_user,
        accion="auth.cambiar_password",
        entidad="Usuario",
        entidad_id=current_user.id,
        detalle={"email": current_user.email, "nombre": current_user.nombre},
        request=request,
    )
    db.commit()

    return {"detalle": "Contraseña actualizada correctamente"}


@app.post("/reservas", response_model=schemas.ReservaOut)
def crear_reserva(
    reserva: schemas.ReservaCreate,
    request: Request,
    current_user: models.Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Crea una reserva para un aula concreta, una fecha y una franja horaria.
    - Profesor: solo puede reservar para sí mismo (se ignora el campo profesor del body).
    - Administrador: puede reservar en nombre de cualquier profesor.
    - Ningún usuario puede reservar en fechas ya pasadas ni en franjas que ya hayan empezado.
    Además, el profesor solo puede reservar en los próximos 7 días.
    """
    ahora = datetime.now(APP_TZ)
    hoy = ahora.date()

    if reserva.fecha < hoy:
        raise HTTPException(
            status_code=400,
            detail="No se puede reservar para fechas pasadas.",
        )

    if reserva.fecha == hoy:
        inicio = hora_inicio_franja(reserva.franja_horaria)
        inicio_dt = datetime.combine(hoy, inicio, tzinfo=APP_TZ)
        if ahora >= inicio_dt:
            raise HTTPException(
                status_code=400,
                detail="No se puede reservar una franja horaria que ya ha empezado o ha pasado.",
            )

    if current_user.rol == schemas.RolEnum.PROFESOR.value:
        max_fecha = hoy + timedelta(days=7)
        if reserva.fecha > max_fecha:
            raise HTTPException(
                status_code=400,
                detail="Un profesor solo puede reservar dentro de los próximos 7 días.",
            )
        profesor_nombre = current_user.nombre
    else:
        # Admin: usa el profesor indicado en el body (o su propio nombre si viene vacío)
        profesor_nombre = reserva.profesor.strip() or current_user.nombre

    solapada = (
        db.query(models.Reserva)
        .filter(
            models.Reserva.fecha == reserva.fecha,
            models.Reserva.aula == reserva.aula.value,
            models.Reserva.franja_horaria == reserva.franja_horaria.value,
        )
        .first()
    )
    if solapada:
        raise HTTPException(
            status_code=400,
            detail="Ya existe una reserva para esa aula, fecha y franja horaria",
        )

    nueva = models.Reserva(
        profesor=profesor_nombre,
        aula=reserva.aula.value,
        fecha=reserva.fecha,
        franja_horaria=reserva.franja_horaria.value,
    )
    db.add(nueva)
    db.flush()
    registrar_auditoria(
        db,
        actor=current_user,
        accion="reserva.crear",
        entidad="Reserva",
        entidad_id=nueva.id,
        detalle={
            "profesor": profesor_nombre,
            "aula": reserva.aula.value,
            "fecha": reserva.fecha.isoformat(),
            "franja_horaria": reserva.franja_horaria.value,
        },
        request=request,
    )
    db.commit()
    db.refresh(nueva)

    enviar_mensaje_teams(
        f"✅ Reserva creada por {profesor_nombre} - {reserva.aula.value}, "
        f"{reserva.franja_horaria.value} el día {reserva.fecha.isoformat()}."
    )

    return schemas.ReservaOut(
        id=nueva.id,
        profesor=nueva.profesor,
        aula=schemas.AulaEnum(nueva.aula),
        fecha=nueva.fecha,
        franja_horaria=schemas.FranjaHorariaEnum(nueva.franja_horaria),
    )


@app.post("/reservas/recurrentes")
def crear_reservas_recurrentes(
    datos: schemas.ReservaRecurrenteCreate,
    request: Request,
    current_user: models.Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    """
    Crea reservas recurrentes (una por semana) hasta el 30 de junio
    del año de la fecha_inicio, manteniendo mismo día de la semana y franja.
    Solo administradores.
    """
    inicio = datos.fecha_inicio
    fin_curso = date(year=inicio.year, month=6, day=30)
    if inicio > fin_curso:
        raise HTTPException(
            status_code=400,
            detail="La fecha de inicio es posterior al fin de curso (30 de junio).",
        )

    profesor_nombre = (datos.profesor or current_user.nombre).strip()
    if not profesor_nombre:
        raise HTTPException(
            status_code=400,
            detail="Debe indicarse un nombre de profesor.",
        )

    creadas = 0
    saltadas: List[date] = []

    dia = inicio
    while dia <= fin_curso:
        solapada = (
            db.query(models.Reserva)
            .filter(
                models.Reserva.fecha == dia,
                models.Reserva.aula == datos.aula.value,
                models.Reserva.franja_horaria == datos.franja_horaria.value,
            )
            .first()
        )
        if solapada:
            saltadas.append(dia)
        else:
            nueva = models.Reserva(
                profesor=profesor_nombre,
                aula=datos.aula.value,
                fecha=dia,
                franja_horaria=datos.franja_horaria.value,
            )
            db.add(nueva)
            creadas += 1

        dia += timedelta(weeks=1)

    db.commit()

    registrar_auditoria(
        db,
        actor=current_user,
        accion="reserva.crear_recurrentes",
        entidad="Reserva",
        entidad_id=None,
        detalle={
            "profesor": profesor_nombre,
            "aula": datos.aula.value,
            "franja_horaria": datos.franja_horaria.value,
            "fecha_inicio": inicio.isoformat(),
            "fecha_fin": fin_curso.isoformat(),
            "creadas": creadas,
            "saltadas": [d.isoformat() for d in saltadas],
        },
        request=request,
    )
    db.commit()

    enviar_mensaje_teams(
        f"📅 Reservas recurrentes creadas por {profesor_nombre} en {datos.aula.value}, "
        f"franja {datos.franja_horaria.value} desde {inicio.isoformat()} hasta {fin_curso.isoformat()}. "
        f"Total nuevas: {creadas}, saltadas por conflicto: {len(saltadas)}."
    )

    return {
        "creadas": creadas,
        "saltadas": [d.isoformat() for d in saltadas],
        "aula": datos.aula.value,
        "franja_horaria": datos.franja_horaria.value,
    }


@app.get("/reservas", response_model=list[schemas.ReservaOut])
def listar_reservas(
    current_user: models.Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Lista las reservas.
    - Administrador: ve todas.
    - Profesor: solo ve sus reservas.
    """
    consulta = db.query(models.Reserva)
    if current_user.rol == schemas.RolEnum.PROFESOR.value:
        consulta = consulta.filter(models.Reserva.profesor == current_user.nombre)

    reservas = consulta.all()

    salida: list[schemas.ReservaOut] = []
    for r in reservas:
        salida.append(
            schemas.ReservaOut(
                id=r.id,
                profesor=r.profesor,
                aula=schemas.AulaEnum(r.aula),
                fecha=r.fecha,
                franja_horaria=schemas.FranjaHorariaEnum(r.franja_horaria),
            )
        )
    return salida


@app.delete("/reservas/{reserva_id}", status_code=204)
def borrar_reserva(
    reserva_id: int,
    request: Request,
    current_user: models.Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    - Profesor: solo puede borrar sus propias reservas.
    - Administrador: puede borrar cualquier reserva.
    """
    reserva = db.query(models.Reserva).get(reserva_id)
    if not reserva:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")

    if current_user.rol == schemas.RolEnum.PROFESOR.value:
        if reserva.profesor != current_user.nombre:
            raise HTTPException(
                status_code=403,
                detail="Un profesor solo puede cancelar sus propias reservas",
            )

    registrar_auditoria(
        db,
        actor=current_user,
        accion="reserva.borrar",
        entidad="Reserva",
        entidad_id=reserva.id,
        detalle={
            "profesor": reserva.profesor,
            "aula": reserva.aula,
            "fecha": reserva.fecha.isoformat(),
            "franja_horaria": reserva.franja_horaria,
        },
        request=request,
    )
    db.delete(reserva)
    db.commit()

    enviar_mensaje_teams(
        f"❌ Reserva cancelada por {current_user.nombre} - {reserva.aula}, "
        f"{reserva.franja_horaria} el día {reserva.fecha.isoformat()} (profesor: {reserva.profesor})."
    )

    return


@app.get("/cuadrante-semanal")
def cuadrante_semanal(
    fecha: date,
    aula: schemas.AulaEnum | None = None,
    current_user: models.Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """
    Devuelve el cuadrante semanal de ocupación.
    - Fila superior: días de la semana (lunes-viernes).
    - Columna izquierda: franjas horarias.
    Cualquier usuario autenticado puede consultarlo.
    """
    # Calculamos lunes-viernes de esa semana y reutilizamos el generador genérico
    lunes = fecha - timedelta(days=fecha.weekday())
    viernes = lunes + timedelta(days=4)
    datos = generar_cuadrante_rango(db, lunes, viernes, aula=aula)
    return {
        "fecha_inicio_semana": datos["fecha_inicio"],
        "fecha_fin_semana": datos["fecha_fin"],
        "dias": datos["dias"],
        "franjas_horarias": datos["franjas_horarias"],
        "aulas": datos["aulas"],
        "cuadrante": datos["cuadrante"],
    }


@app.get("/estadisticas/reservas")
def estadisticas_reservas(
    admin: models.Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """
    Estadísticas globales de reservas hasta la fecha actual.
    Incluye totales y desglose por aula.
    Solo administradores.
    """
    hoy = date.today()

    total = db.query(func.count(models.Reserva.id)).scalar() or 0

    por_aula_rows = (
        db.query(models.Reserva.aula, func.count(models.Reserva.id))
        .filter(models.Reserva.fecha <= hoy)
        .group_by(models.Reserva.aula)
        .all()
    )
    por_aula = {aula: count for aula, count in por_aula_rows}

    return {
        "hasta_fecha": hoy.isoformat(),
        "total_reservas": total,
        "por_aula": por_aula,
    }


@app.get("/estadisticas/aula/{aula}")
def estadisticas_por_aula(
    aula: schemas.AulaEnum,
    admin: models.Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """
    Estadísticas detalladas para un aula concreta hasta la fecha actual.
    Solo administradores.
    """
    hoy = date.today()

    total = (
        db.query(func.count(models.Reserva.id))
        .filter(models.Reserva.aula == aula.value, models.Reserva.fecha <= hoy)
        .scalar()
        or 0
    )

    por_profesor_rows = (
        db.query(models.Reserva.profesor, func.count(models.Reserva.id))
        .filter(models.Reserva.aula == aula.value, models.Reserva.fecha <= hoy)
        .group_by(models.Reserva.profesor)
        .order_by(func.count(models.Reserva.id).desc())
        .all()
    )
    por_profesor = {prof: count for prof, count in por_profesor_rows}

    return {
        "hasta_fecha": hoy.isoformat(),
        "aula": aula.value,
        "total_reservas": total,
        "por_profesor": por_profesor,
    }


def _generar_excel_cuadrante(datos: Dict[str, Any]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    dias = datos["dias"]
    franjas = datos["franjas_horarias"]
    aulas = datos["aulas"]
    cuadrante = datos["cuadrante"]

    for aula in aulas:
        ws = wb.create_sheet(title=aula[:31])
        ws.append(["Franja horaria"] + dias)
        for franja in franjas:
            fila = [franja]
            for dia in dias:
                info = cuadrante[aula][franja][dia]
                fila.append(info["profesor"] if info else "")
            ws.append(fila)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _generar_pdf_cuadrante(datos: Dict[str, Any]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elementos = []
    styles = getSampleStyleSheet()

    dias = datos["dias"]
    franjas = datos["franjas_horarias"]
    aulas = datos["aulas"]
    cuadrante = datos["cuadrante"]

    for idx, aula in enumerate(aulas):
        if idx > 0:
            elementos.append(Spacer(1, 12))

        titulo = Paragraph(f"Cuadrante - {aula}", styles["Heading2"])
        elementos.append(titulo)
        elementos.append(Spacer(1, 6))

        encabezados = ["Franja horaria"] + dias
        data = [encabezados]
        for franja in franjas:
            fila = [franja]
            for dia in dias:
                info = cuadrante[aula][franja][dia]
                fila.append(info["profesor"] if info else "")
            data.append(fila)

        tabla = Table(data, repeatRows=1)
        tabla.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ]
            )
        )
        elementos.append(tabla)

    doc.build(elementos)
    return buffer.getvalue()


def enviar_informe_semanal_ocupacion() -> None:
    """
    Genera y envía al administrador el informe semanal de ocupación
    (semana anterior, lunes-viernes) en Excel y PDF.
    """
    if not SMTP_HOST or not SMTP_FROM:
        return

    db = SessionLocal()
    try:
        hoy = datetime.now(APP_TZ).date()
        lunes_actual = hoy - timedelta(days=hoy.weekday())
        lunes_anterior = lunes_actual - timedelta(weeks=1)
        viernes_anterior = lunes_anterior + timedelta(days=4)

        datos = generar_cuadrante_rango(db, lunes_anterior, viernes_anterior, aula=None)

        excel_bytes = _generar_excel_cuadrante(datos)
        pdf_bytes = _generar_pdf_cuadrante(datos)

        admins = (
            db.query(models.Usuario)
            .filter(models.Usuario.rol == schemas.RolEnum.ADMIN.value, models.Usuario.activo.is_(True))
            .all()
        )
        destinatarios = [u.email for u in admins if u.email]
        if not destinatarios:
            return

        asunto = "Informe semanal de ocupación de aulas"
        cuerpo = (
            "Se adjuntan los cuadrantes de ocupación de aulas de la semana anterior "
            f"({lunes_anterior.isoformat()} a {viernes_anterior.isoformat()})."
        )

        adjuntos = [
            ("ocupacion_semana_anterior.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            ("ocupacion_semana_anterior.pdf", pdf_bytes, "application/pdf"),
        ]

        enviar_correo(asunto, cuerpo, destinatarios, adjuntos)
    finally:
        db.close()


@app.on_event("startup")
def configurar_scheduler() -> None:
    """
    Configura un scheduler en segundo plano para enviar el informe los lunes a la 01:00.
    """
    global scheduler
    if scheduler is not None:
        return

    scheduler_local = AsyncIOScheduler(timezone=APP_TZ)
    scheduler_local.add_job(
        enviar_informe_semanal_ocupacion,
        CronTrigger(day_of_week="mon", hour=1, minute=0),
        id="informe_semanal_ocupacion",
        replace_existing=True,
    )
    scheduler_local.start()
    scheduler = scheduler_local


